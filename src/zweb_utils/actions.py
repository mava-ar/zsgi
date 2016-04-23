import openpyxl
from django.http import HttpResponse


def export_xlsx(modeladmin, request, queryset):

    from openpyxl.cell import get_column_letter
    opts = modeladmin.model._meta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format("{}".format(opts).replace('.', '_'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "{}".format(opts).replace('.', '_')

    field_names = [field.name for field in opts.fields]
    row_num = 0

    for col_num in range(len(field_names)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = field_names[col_num]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = 35

    for obj in queryset:
        row_num += 1
        try:
            for col_num in range(len(field_names)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = getattr(obj, field_names[col_num], '')
                c.style.alignment.wrap_text = True
        except Exception as e:
            print(e)

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"
